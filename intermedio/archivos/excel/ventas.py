# from openpyxl import load_workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.utils import get_column_letter
# from pathlib import Path

# # Ruta del archivo Excel
# ruta_archivo = Path(__file__).parent / "reporte.xlsx"
# wb = load_workbook(ruta_archivo)
# hoja = wb["Ventas"]

# # Validar que haya datos (mínimo encabezados + 1 fila)
# if hoja.max_row < 2:
#     raise SystemExit(
#         "⚠️ No hay datos suficientes para crear una tabla en la hoja Ventas.")

# nombre_tabla = "TablaVentas"

# # ✅ Eliminar tabla previa si ya existía (forma segura)
# if nombre_tabla in hoja.tables:
#     del hoja.tables[nombre_tabla]

# # ✅ Detectar rango de datos dinámicamente
# ultima_fila = hoja.max_row
# ultima_columna = hoja.max_column
# columna_final = get_column_letter(ultima_columna)
# rango = f"A1:{columna_final}{ultima_fila}"

# # ✅ Crear tabla nueva
# tabla = Table(displayName=nombre_tabla, ref=rango)

# # ✅ Estilo visual
# estilo = TableStyleInfo(
#     name="TableStyleMedium9",
#     showFirstColumn=False,
#     showLastColumn=False,
#     showRowStripes=True,
#     showColumnStripes=False
# )
# tabla.tableStyleInfo = estilo

# # ✅ Agregar a hoja
# hoja.add_table(tabla)

# # ✅ Guardar
# wb.save(ruta_archivo)
# print("✅ Tabla creada exitosamente en la hoja 'Ventas'")


from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from pathlib import Path

# Ruta del archivo Excel
ruta_archivo = Path(__file__).parent / "reporte.xlsx"
wb = load_workbook(ruta_archivo)
hoja = wb["Ventas"]

# ✅ Ajustar ancho de columnas automáticamente
for columna in hoja.columns:
    max_length = 0
    col = columna[0].column_letter
    for cell in columna:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    hoja.column_dimensions[col].width = max_length + 2

# ✅ Dar estilo a encabezados
for cell in hoja[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="4F81BD")

# ✅ Alinear celdas al centro excepto productos (texto)
for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=5):
    fila[0].alignment = Alignment(horizontal="center")
    fila[1].alignment = Alignment(horizontal="left")
    fila[2].alignment = Alignment(horizontal="right")
    fila[3].alignment = Alignment(horizontal="center")
    fila[4].alignment = Alignment(horizontal="right")

# ✅ Formato moneda en columnas C y E
for fila in hoja.iter_rows(min_row=2, min_col=3, max_col=5):
    fila[0].number_format = '"$"#,##0'
    fila[2].number_format = '"$"#,##0'

# ✅ Agregar bordes
thin = Side(border_style="thin", color="000000")
for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=5):
    for celda in fila:
        celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)

wb.save(ruta_archivo)
print("✅ Formato aplicado correctamente a la hoja 'Ventas'")
