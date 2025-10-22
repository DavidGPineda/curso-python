from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from pathlib import Path

# Ruta al archivo
ruta_archivo = Path(__file__).parent / "reporte.xlsx"
wb = load_workbook(ruta_archivo)
hoja = wb["Inicio"]

# Limpiar contenido anterior por si se ejecuta varias veces
hoja.delete_rows(1, hoja.max_row)

# Combinar celdas para el título
hoja.merge_cells("A1:D1")
hoja["A1"] = "Reporte General"
hoja["A1"].font = Font(size=16, bold=True)  # Negrita + grande
hoja["A1"].alignment = Alignment(horizontal="center")  # Centrar texto
hoja["A1"].fill = PatternFill(
    start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Fondo azul

# Subtitulo
hoja.merge_cells("A2:D2")
hoja["A2"] = "Generado con Python"
hoja["A2"].font = Font(size=12, italic=True, color="000000")
hoja["A2"].alignment = Alignment(horizontal="left")

hoja["A3"] = "Hola"

# Fecha automatica
hoja.merge_cells("A4:B4")
fecha_actual = datetime.now().strftime("%d/%m/%Y")
hoja["A4"] = f"Fecha de generación: {fecha_actual}"
hoja["A4"].font = Font(size=11)

# Ajustar ancho de columna
for col in ["A", "B", "C", "D"]:
    hoja.column_dimensions[col].width = 25

# Guardar cambios
wb.save(ruta_archivo)
print("Hoja 'Inicio' formateada correctamente con estilo")
