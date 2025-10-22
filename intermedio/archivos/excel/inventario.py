from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path


ruta_archivo = Path(__file__).parent / "reporte.xlsx"
wb = load_workbook(ruta_archivo)
hoja = wb["Inventario"]

# Encabezados
encabezados = ["ID", "Producto", "Categoria",
               "Talla", "Stock", "Precio", "Valor Total"]
hoja.append(encabezados)

# Insertar datos de ejemplo (10 productos)
productos = [
    [1, "Camisa básica", "Hombre", "M", 25, 25000],
    [2, "Jean clásico", "Hombre", "32", 15, 80000],
    [3, "Sudadera deportiva", "Unisex", "L", 18, 95000],
    [4, "Chaqueta impermeable", "Mujer", "S", 12, 150000],
    [5, "Pantaloneta", "Hombre", "M", 30, 35000],
    [6, "Blusa manga larga", "Mujer", "M", 20, 45000],
    [7, "Vestido casual", "Mujer", "M", 10, 120000],
    [8, "Zapatos deportivos", "Unisex", 41, 14, 180000],
    [9, "Gorra ajustable", "Unisex", "-", 40, 30000],
    [10, "Medias deportivas (x3)", "Unisex", "-", 50, 20000]
]

for fila in productos:
    # Valor total
    fila.append(f"=E{productos.index(fila)+2}*F{productos.index(fila)+2}")
    hoja.append(fila)

# Estilo encabezados
for celda in hoja[1]:
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="4F81BD")
    celda.alignment = Alignment(horizontal="center", vertical="center")

# Ajuste de ancho automatico
for columna in hoja.columns:
    max_len = 0
    col = columna[0].column_letter
    for cell in columna:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
        hoja.column_dimensions[col].width = max_len + 2

# Formato numerico
for fila in hoja.iter_rows(min_row=2, min_col=5, max_col=6):
    fila[0].alignment = Alignment(horizontal="center")
    fila[0].number_format = "0"
    fila[1].number_format = '"$"#,##0'

# Bordes
thin = Side(border_style="thin", color="000000")
for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=7):
    for celda in fila:
        celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)

wb.save(ruta_archivo)
print("Inventario creado correctamente con formato profesional")
