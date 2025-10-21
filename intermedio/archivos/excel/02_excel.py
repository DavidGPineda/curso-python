from openpyxl import load_workbook, Workbook
from pathlib import Path

ruta_archivo = Path(__file__).parent / "empleados.xlsx"

# cargar el archivo de excel
# wb = load_workbook(ruta_archivo)
# wb = load_workbook("empleados.xlsx")

# seleccionar  la hoja (Sheet1 por defecto)
# hoja = wb.active  # o wb["Sheet1"]


# Leer celdas especificas
# print(hoja["A1"].value)  # Primera celda (encabezado)
# print(hoja["A2"].value)  # Primer nombre
# print(hoja["B2"].value)  # Primera edad
# print(hoja["C2"].value)  # Primera ciudad
# print(hoja["D2"].value)  # Primer salario


# Paso 3: Recorrer todo el archivo Excel

# Recorrer todas las filas
# for fila in hoja.iter_rows(values_only=True):
#     print(fila)

# ciudad_buscar = input("Ingresa una ciudad: ").strip()

# encontrados = False

# print(f"Empleados de {ciudad_buscar}:\n")
# for fila in hoja.iter_rows(min_row=2, values_only=True):
#     # min_row=2 para saltar encabezado
#     nombre, edad, ciudad, salario = fila
#     # Desempaquetar cada fila
#     if ciudad.lower() == ciudad_buscar.lower():
#         print(f"{nombre} - {edad} años - ${salario}")
#         encontrados = True

# if not encontrados:
#     print("No hay empleados en esa ciudad")

# escribir en Excel con Python ✍️📊
# ✏️ Paso 1: Insertar una nueva fila

# print("=== Agregar nuevo empleado ===")
# nombre = input("Nombre completo: ")
# edad = input("Edad: ")
# ciudad = input("Ciudad: ")
# salario = input("Salario: ")

# Datos a agregar
# nuevo_empleado = ["María Sánchez", 27, "Medellín", 3800000]

# Lista de nuevas filas (cada sublista es un empleado)
# nuevos_empleados = [
#     ["Paola Herrera", 32, "Bogotá", 4500000],
#     ["Diego Ramírez", 29, "Medellín", 3700000],
#     ["Lucía Gómez", 26, "Cali", 3300000],
# ]

# Agregar todas las filas
# for empleado in nuevos_empleados:
#     hoja.append(empleado)


# Agregar fila al final
# hoja.append([nombre, edad, ciudad, salario])

# Guardar cambios
# wb.save("empleados.xlsx")

# print("Empleado agregado correctamente")

# Código básico para actualizar un dato

# Buscar y actualizar salario
# nombre_buscar = input("Ingrese el nombre del empleado a actualizar: ").strip()
# nuevo_salario = int(input("Ingrese el nuevo salario: "))

# encontrado = False

# for fila in hoja.iter_rows(min_row=2):
#     nombre = fila[0].value  # columna A = nombre
#     if nombre.lower() == nombre_buscar.lower():
#         fila[3].value = nuevo_salario  # columna D = salario
#         encontrado = True
#         break

# if encontrado:
#     wb.save("empleados.xlsx")
#     print("Salario actualizado correctamente")
# else:
#     print("No se encontro al empleado")


# Borrar una fila

# Pedir nombre para eliminar
# nombre_buscar = input("Ingrese el nombre del empleado a eliminar: ").strip()

# fila_eliminar = None

# Buscar fila
# for fila in hoja.iter_rows(min_row=2):
#     if fila[0].value and fila[0].value.lower() == nombre_buscar.lower():
#         fila_eliminar = fila[0].row  # Guardamos el número de la fila real
#         break

# Eliminar si se encontró
# if fila_eliminar:
#     hoja.delete_rows(fila_eliminar, 1)  # Borrar una fila
#     wb.save("empleados.xlsx")
#     print(f"Empleado '{nombre_buscar}' eliminado correctamente.")
# else:
#     print("Empleado no encontrado")

# Mostrar nombres de las hojas
# print("Hojas del archivo Excel:")
# for hoja in wb.sheetnames:
#     print(" -", hoja)

# --------------------------------------
