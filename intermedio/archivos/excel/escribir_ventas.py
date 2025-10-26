from openpyxl import load_workbook
from pathlib import Path

ruta_archivo = Path(__file__).parent / "reportes.xlsx"
wb = load_workbook(ruta_archivo)
hoja = wb["Ventas"]

# Limpiar hoja completamente
hoja.delete_rows(1, hoja.max_row)

# Escribir encabezados en la primera fila
hoja["A1"] = "ID"
hoja["B1"] = "Producto"
hoja["C1"] = "Precio"
hoja["D1"] = "Cantidad"
hoja["E1"] = "Total"

# Datos
ventas = [
    [1, "Laptop Dell", 3500000, 2],
    [2, "Mouse Logitech", 80000, 5],
    [3, "Monitor Samsung", 900000, 1],
    [4, "Teclado Redragon", 120000, 3]
]

# Agregar filas desde A2 en adelante
for i, fila in enumerate(ventas, start=2):
    hoja[f"A{i}"] = fila[0]
    hoja[f"B{i}"] = fila[1]
    hoja[f"C{i}"] = fila[2]
    hoja[f"D{i}"] = fila[3]
    hoja[f"E{i}"] = fila[2] * fila[3]

wb.save(ruta_archivo)
print("✅ Encabezados y datos escritos correctamente desde A1")
