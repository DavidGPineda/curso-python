from openpyxl import load_workbook
from pathlib import Path

ruta_archivo = Path(__file__).parent / "reporte.xlsx"
wb = load_workbook(ruta_archivo)

# Orden deseado
orden = ["Inicio", "Ventas", "Inventario", "Resumen"]

# Reordear solo si todas las hojas existen
if all(hoja in wb.sheetnames for hoja in orden):
    for index, hoja in enumerate(orden):
        wb.move_sheet(wb[hoja], offset=index - wb.sheetnames.index(hoja))
    wb.save(ruta_archivo)
    print("✅ Hojas reordenadas correctamente")
else:
    print("No se pudo reordenar porque falta alguna hoja necesaria")
