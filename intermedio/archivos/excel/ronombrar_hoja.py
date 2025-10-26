from openpyxl import load_workbook
from pathlib import Path

# Ruta al archivo Excel
ruta_archivo = Path(__file__).parent / "reportes.xlsx"

# Abrir archivo
wb = load_workbook(ruta_archivo)

# Renombrar hoja si exise "Sheet"
if "Sheet" in wb.sheetnames:
    hoja = wb["Sheet"]
    hoja.title = "Inicio"
    wb.save(ruta_archivo)
    print("Hoja 'Sheet' renombrada a 'Inicio'")
else:
    print("No existe una hoja llamada 'Sheet' en el archivo")
