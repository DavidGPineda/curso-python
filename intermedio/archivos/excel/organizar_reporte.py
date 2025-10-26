from openpyxl import load_workbook, Workbook
from pathlib import Path

ruta_archivo = Path(__file__).parent / "reportes.xlsx"

wb = load_workbook(ruta_archivo)

# Paso 1: Crear un archivo Excel nuevo
# wb = Workbook()

# Guardarlo en la carpeta actual
# wb.save(ruta_archivo)

# print("Archivo roporte.xlsx creado correctamente")


# Paso 2: Crear nuevas hojas

# Verificar si existe

# if not ruta_archivo:
#     print(f"No se encontro {ruta_archivo}, Primero crea el archivo")
#     exit()

# # Abrimos el archivo que creamos
# wb = load_workbook(ruta_archivo)

# # Crearmos nuevas hojas
# wb.create_sheet("Ventas")
# wb.create_sheet("Inventario")
# wb.create_sheet("Resumen")

# # # Guardar cambios
# wb.save(ruta_archivo)

# print("Hojas creadas correctamente")

# Paso: Listar las hojas del workbook
print("Hojas en reporte.xlsx")
for hoja in wb.sheetnames:
    print(" -", hoja)
